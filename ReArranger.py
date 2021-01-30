import os
import sys

import pandas as pd
import xlrd

from openpyxl import load_workbook
from PyQt5 import QtWidgets as qtw
from PyQt5 import QtCore as qtc



class MainWindow(qtw.QWidget):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Your code will go here

        cur_folder_label = 'Please browse for directory with XLS, XLSX or CSV files.'
        self.directory = ''
        self.col_list = []
    ## setting up the objects

        self.browse_button = qtw.QPushButton('1. Browse')
        self.master_file_select_button = qtw.QPushButton('2. Select Master File')
        self.go_button = qtw.QPushButton('3. Process Files')
        self.go_button.setGeometry(400,400,100,100)
        self.quit_button = qtw.QPushButton('4. Quit')
        self.cur_folder_label = qtw.QLabel(f'Current Folder: \n{cur_folder_label}')
        self.master_file_label = qtw.QLabel(f'Master file : None Chosen')
        self.blank_spacer = qtw.QLabel('Master Column List (By Order)')
        self.master_column_list = qtw.QListWidget()



        layout = qtw.QGridLayout()
        layout.addWidget(self.cur_folder_label,0,0)
        layout.addWidget(self.browse_button,2,0)
        layout.addWidget(self.go_button,6,3)
        layout.addWidget(self.quit_button,6,3,alignment=qtc.Qt.AlignBottom)
        layout.addWidget(self.master_file_select_button,3,0)
        layout.addWidget(self.master_file_label,4,0)
        layout.addWidget(self.blank_spacer,5,0)
        layout.addWidget(self.master_column_list,6,0)


        self.browse_button.clicked.connect(self.select_folder)
        self.go_button.clicked.connect((self.lets_do_this))
        self.master_file_select_button.clicked.connect(self.select_master_file)
        self.quit_button.clicked.connect(self.close)
        self.setLayout(layout)

        # Your code ends here
        self.show()


    def select_master_file(self):
        valid_files = ['xls', 'XLS', 'xlsx', 'XLSX', 'csv', 'CSV']
        self.master_file_path = qtw.QFileDialog.getOpenFileName(self, 'Please select the Master File')
        if self.master_file_path == ('', ''):
            self.warning = qtw.QMessageBox.information(self, 'Attention','Master File was NOT selected')

        else:
            self.master_column_list.clear()
            self.master_fileName = self.master_file_path[0].split('/')[-1]
            if self.master_fileName.split('.')[1] not in valid_files:
                self.warning = qtw.QMessageBox.critical(self, 'Error',f'{self.master_fileName} is not a valid file')
            else:
                self.master_file_label.setText(f'Master file : {self.master_fileName}')
                self.col_list = get_master_column_list(self.master_file_path[0])
                for col in self.col_list:
                    self.master_column_list.insertItem(1, str(col))


    def select_folder(self):
        self.directory = qtw.QFileDialog.getExistingDirectory(self, "Select Directory")
        if not self.directory:
            self.warning = qtw.QMessageBox.information(self, 'Attention','Input Directory not selected')
        else:
            search_for_files(self.directory)
            if len(search_for_files(self.directory)) == 0:
                self.warning = qtw.QMessageBox.information(self, 'Attention', 'No valid files were found')
            self.cur_folder_label.setText(f'Current Path :\n'
                                          f'{self.directory}\n'
                                          f'Found {len(search_for_files(self.directory))} file(s).')

    def lets_do_this(self):
        if self.directory == '':
            self.warning = qtw.QMessageBox.information(self, 'Attention', f'Please Choose Input Directory.\n (Step 1)')
        elif not self.col_list:
            self.warning = qtw.QMessageBox.information(self, 'Attention', 'Please Choose Master File.\n (Step 2)')

        else:
            self.warning = qtw.QMessageBox.information(self, 'Attention', 'Please Select Output Folder.')
            output_directory = qtw.QFileDialog.getExistingDirectory(self, "Select Directory")
            if not output_directory:
                self.warning = qtw.QMessageBox.information(self, 'Attention', f'Operation cancelled by user.')
            else:
                files_to_process = search_for_files(self.directory)
                for i in files_to_process:
                    if len(i.split('.')) == 1: # Ignoring directories and/or files an extention
                        pass
                    elif i.split('.')[-1] == 'xls':  # Process xls
                        df = pd.read_excel(self.directory + '/' + i) # Making a DataFrame
                        df2 = df.reindex(columns=self.col_list)
                        writer = pd.ExcelWriter(f'{output_directory}/{i}x', engine='xlsxwriter')
                        df2.to_excel(writer, index=False)
                        writer.save()

                    if i.split('.')[-1] == 'xlsx':  # Process xlsx
                        df = pd.read_excel(self.directory + '/' + i) # Making a DataFrame
                        df2 = df.reindex(columns=self.col_list)
                        writer = pd.ExcelWriter(output_directory + '/' + i, engine='xlsxwriter')
                        df2.to_excel(writer, index=False)
                        writer.save()
                    elif i.split('.')[-1] == 'csv':  # Process csv
                        df = pd.read_csv(self.directory + '/' + i) # Making a DataFrame
                        df2 = df.reindex(columns=self.col_list)
                        df2.to_csv(output_directory + '/' + i, index = False)
                    else:
                        pass
            self.warning = qtw.QMessageBox.information(self, 'Attention', f'Processed {len(search_for_files(self.directory))} files.')


def get_master_column_list(file):
    master_column_list = []
    if file.split('.')[1] == 'xlsx':
        wb = load_workbook(file)
        ws = wb.active
        for cell in ws[1]:
            master_column_list.append(cell.value)

    elif file.split('.')[1] == 'xls':
        wb = xlrd.open_workbook(file)
        for cell in wb[0][0]:
            if cell.value == 0.0:
                cell.value = 0
            master_column_list.append(cell.value)

    elif file.split('.')[1] == 'csv':
        with open(file, 'r') as wb:
            firstLine = wb.readline()
            wb = firstLine.replace('"','').replace('\n','').strip().split(',')
            for cell in wb:
                master_column_list.append(cell.strip())

    return master_column_list

def search_for_files(current_dir):
    matching_files = []
    all_files = [file for file in os.listdir(current_dir) if not file.startswith('.')]
    for i in all_files:
        if len(i.split('.')) == 1:
            pass
        elif i.split('.')[1] == 'xlsx':
            matching_files.append(i)
        elif i.split('.')[1] == 'xls':
            matching_files.append(i)
        elif i.split('.')[1] == 'csv':
            matching_files.append(i)
        else:
            pass
    return matching_files







if __name__ == '__main__':
    app = qtw.QApplication(sys.argv)
    style = """
            QWidget{
                background: #262D37;
            }
            QLabel{
                color: #fff;
            }
            QPushButton
            {
                color: white;
                background: #0577a8;
                border: 1px #DADADA solid;
                padding: 5px 10px;
                border-radius: 2px;
                font-weight: bold;
                font-size: 9pt;
                outline: none;
            }
            QPushButton:hover{
                border: 1px #C6C6C6 solid;
                color: #fff;
                background: #0892D0;
            }
            QListWidget {
                padding: 1px;
                color: #fff;
                border-style: solid;
                border: 2px solid #fff;
                border-radius: 8px;
            }
        """
    app.setStyleSheet(style)
    w = MainWindow(windowTitle='Excel / CSV Column Rearranger')
    sys.exit(app.exec_())
